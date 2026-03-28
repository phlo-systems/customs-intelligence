"""
bulk_country_loader.py — Load tariff data for multiple countries from Excel/CSV sources.

Handles: Uruguay, Chile, UAE, Saudi Arabia, Oman, Argentina + future countries.

Usage:
    export $(grep -v '^#' .env | xargs)
    python3 -m tariff_parser.bulk_country_loader --country UY
    python3 -m tariff_parser.bulk_country_loader --country CL
    python3 -m tariff_parser.bulk_country_loader --country AE  # also loads SA, OM
    python3 -m tariff_parser.bulk_country_loader --country AR  # copies from BR
    python3 -m tariff_parser.bulk_country_loader --all
"""

import argparse
import logging
import os
import re
import sys
from datetime import date, datetime, timezone

import openpyxl
import requests

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s — %(message)s")
logger = logging.getLogger("bulk_loader")


class BulkCountryLoader:
    def __init__(self, supabase_url: str, supabase_key: str):
        self.url = supabase_url.rstrip("/")
        self.key = supabase_key
        self.headers = {
            "apikey": self.key, "Authorization": f"Bearer {self.key}",
            "Content-Type": "application/json",
            "Prefer": "resolution=merge-duplicates,return=minimal",
        }
        self.today = date.today().isoformat()

    def _upsert(self, table, records, batch_size=200):
        total = 0
        for i in range(0, len(records), batch_size):
            chunk = records[i:i + batch_size]
            resp = requests.post(f"{self.url}/rest/v1/{table}", headers=self.headers, json=chunk, timeout=30)
            if resp.status_code in (200, 201):
                total += len(chunk)
            else:
                logger.error("Upsert %s batch %d failed: %s", table, i, resp.text[:150])
        return total

    def _ensure_country(self, code, code3, name, region, currency, valuation="CIF"):
        requests.post(f"{self.url}/rest/v1/country", headers=self.headers, json=[{
            "countrycode": code, "countrycode3": code3, "countryname": name,
            "region": region, "currencycode": currency, "valuationbasis": valuation,
        }], timeout=10)

    def _update_freshness(self, country, rows, source):
        now = datetime.now(timezone.utc).isoformat()
        requests.post(f"{self.url}/rest/v1/data_freshness", headers=self.headers, json=[{
            "countrycode": country, "datatype": "BCD_RATES",
            "lastsyncat": now, "rowcount": rows,
            "sourcename": source, "staleafterhours": 720,
        }], timeout=10)

    # ═════════════════════════════════════════════════════════════════════
    # URUGUAY
    # ═════════════════════════════════════════════════════════════════════
    def load_uruguay(self, filepath="/tmp/uy_10digit.xlsx"):
        self._ensure_country("UY", "URY", "Uruguay", "South America", "UYU")
        logger.info("Loading Uruguay from %s", filepath)

        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb[wb.sheetnames[0]]

        comm, mfn = [], []
        for row in ws.iter_rows(min_row=3, values_only=True):
            code_raw = str(row[0] or "").strip()
            desc = str(row[1] or "").strip()
            aec_rate = row[2]

            # Clean code
            code = code_raw.replace(".", "").replace(" ", "")
            if len(code) < 8 or not code[:4].isdigit():
                continue
            # Take first 8 digits
            code8 = code[:8]

            rate = float(aec_rate) if aec_rate is not None else None
            desc_clean = re.sub(r"^\s*[-–—]+\s*", "", desc)[:500]

            comm.append({
                "commoditycode": code8, "countrycode": "UY",
                "subheadingcode": code8[:6], "hsversion": "HS 2022",
                "nationaldescription": desc_clean, "codelength": "8-digit", "isactive": True,
            })
            if rate is not None:
                mfn.append({
                    "commoditycode": code8, "countrycode": "UY",
                    "ratecategory": "APPLIED", "dutybasistype": "AD_VALOREM",
                    "appliedmfnrate": rate, "dutyexpression": f"{rate}%",
                    "valuationbasis": "CIF",
                    "effectivefrom": self.today, "effectiveto": None,
                })
        wb.close()

        # Deduplicate
        seen = set()
        comm_dedup = [c for c in comm if c["commoditycode"] not in seen and not seen.add(c["commoditycode"])]
        mfn_dedup = {m["commoditycode"]: m for m in mfn}

        c = self._upsert("commodity_code", comm_dedup)
        m = self._upsert("mfn_rate", list(mfn_dedup.values()))
        self._update_freshness("UY", c, "MEF Arancel Nacional Excel (Feb 2025)")
        logger.info("UY: %d commodity codes, %d MFN rates", c, m)
        return c

    # ═════════════════════════════════════════════════════════════════════
    # CHILE
    # ═════════════════════════════════════════════════════════════════════
    def load_chile(self, filepath="/tmp/cl_items.xlsx"):
        self._ensure_country("CL", "CHL", "Chile", "South America", "CLP")
        logger.info("Loading Chile from %s", filepath)

        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb[wb.sheetnames[0]]

        comm, mfn = [], []
        for row in ws.iter_rows(min_row=2, values_only=True):
            code_raw = str(row[0] or "").strip()
            desc = str(row[1] or "").strip()
            rate = row[3]  # Column D = Ad valorem rate

            code = code_raw.replace(".", "").replace(" ", "")
            if len(code) < 8 or not code[:4].isdigit():
                continue

            desc_clean = re.sub(r"^\s*[-–—]+\s*", "", desc)[:500]
            try:
                rate_val = float(rate) if rate is not None else 6.0
            except (ValueError, TypeError):
                rate_val = 6.0  # Chile default 6% uniform tariff

            comm.append({
                "commoditycode": code, "countrycode": "CL",
                "subheadingcode": code[:6], "hsversion": "HS 2022",
                "nationaldescription": desc_clean, "codelength": "8-digit", "isactive": True,
            })
            mfn.append({
                "commoditycode": code, "countrycode": "CL",
                "ratecategory": "APPLIED", "dutybasistype": "AD_VALOREM",
                "appliedmfnrate": rate_val, "dutyexpression": f"{rate_val}%",
                "valuationbasis": "CIF",
                "effectivefrom": self.today, "effectiveto": None,
            })
        wb.close()

        seen = set()
        comm_dedup = [c for c in comm if c["commoditycode"] not in seen and not seen.add(c["commoditycode"])]
        mfn_dedup = {m["commoditycode"]: m for m in mfn}

        c = self._upsert("commodity_code", comm_dedup)
        m = self._upsert("mfn_rate", list(mfn_dedup.values()))
        self._update_freshness("CL", c, "Aduana Chile Arancel Excel (2023)")
        logger.info("CL: %d commodity codes, %d MFN rates", c, m)
        return c

    # ═════════════════════════════════════════════════════════════════════
    # UAE + SAUDI + OMAN (GCC Common External Tariff)
    # ═════════════════════════════════════════════════════════════════════
    def load_gcc(self, filepath="/tmp/ae_hscode.xlsx"):
        self._ensure_country("AE", "ARE", "United Arab Emirates", "Middle East", "AED")
        self._ensure_country("SA", "SAU", "Saudi Arabia", "Middle East", "SAR")
        self._ensure_country("OM", "OMN", "Oman", "Middle East", "OMR")
        logger.info("Loading GCC (AE/SA/OM) from %s", filepath)

        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb[wb.sheetnames[0]]

        entries = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            old_code = str(row[0] or "").strip()
            desc_en = str(row[3] or "").strip()
            duty_pct = row[5]

            if len(old_code) < 8 or not old_code[:4].isdigit():
                continue

            code8 = old_code[:8]
            rate = float(duty_pct) if duty_pct is not None else 5.0  # GCC default 5%
            desc_clean = re.sub(r"^\s*[-–—]+\s*", "", desc_en)[:500]

            entries.append({"code": code8, "desc": desc_clean, "rate": rate})
        wb.close()

        # Deduplicate by 8-digit code
        seen = {}
        for e in entries:
            if e["code"] not in seen:
                seen[e["code"]] = e

        unique = list(seen.values())
        logger.info("GCC: %d unique 8-digit codes", len(unique))

        # Load for all 3 GCC countries
        vat_rates = {"AE": 5.0, "SA": 15.0, "OM": 5.0}

        for country in ["AE", "SA", "OM"]:
            comm, mfn, vat = [], [], []
            for e in unique:
                comm.append({
                    "commoditycode": e["code"], "countrycode": country,
                    "subheadingcode": e["code"][:6], "hsversion": "HS 2022",
                    "nationaldescription": e["desc"], "codelength": "8-digit", "isactive": True,
                })
                mfn.append({
                    "commoditycode": e["code"], "countrycode": country,
                    "ratecategory": "APPLIED", "dutybasistype": "AD_VALOREM",
                    "appliedmfnrate": e["rate"], "dutyexpression": f"{e['rate']}%",
                    "valuationbasis": "CIF",
                    "effectivefrom": self.today, "effectiveto": None,
                })
                vat.append({
                    "commoditycode": e["code"], "countrycode": country,
                    "taxtype": "VAT", "taxcategory": "STANDARD",
                    "rate": vat_rates[country],
                    "vatbasis": "CUSTOMS_VALUE_PLUS_DUTY",
                    "postponedaccounting": False, "reliefavailable": False,
                    "effectivefrom": self.today, "effectiveto": None,
                    "notes": None,
                })

            c = self._upsert("commodity_code", comm)
            m = self._upsert("mfn_rate", mfn)
            v = self._upsert("vat_rate", vat)
            self._update_freshness(country, c, "Dubai Customs HSCodeMaster Excel (GCC CET 2025)")
            logger.info("%s: %d codes, %d MFN rates, %d VAT rates", country, c, m, v)

        return len(unique)

    # ═════════════════════════════════════════════════════════════════════
    # ARGENTINA (copy from Brazil — MERCOSUR shared TEC)
    # ═════════════════════════════════════════════════════════════════════
    def load_argentina_from_brazil(self):
        self._ensure_country("AR", "ARG", "Argentina", "South America", "ARS")
        logger.info("Loading Argentina from Brazil's NCM data (MERCOSUR shared TEC)")

        # Fetch BR commodity codes
        all_br = []
        offset = 0
        while True:
            resp = requests.get(
                f"{self.url}/rest/v1/commodity_code?countrycode=eq.BR&isactive=eq.true"
                f"&select=commoditycode,subheadingcode,nationaldescription"
                f"&offset={offset}&limit=1000",
                headers={**self.headers, "Prefer": ""}, timeout=15)
            batch = resp.json()
            if not batch: break
            all_br.extend(batch)
            offset += len(batch)
            if len(batch) < 1000: break

        # Fetch BR MFN rates
        all_br_mfn = []
        offset = 0
        while True:
            resp = requests.get(
                f"{self.url}/rest/v1/mfn_rate?countrycode=eq.BR&effectiveto=is.null"
                f"&select=commoditycode,appliedmfnrate,dutyexpression"
                f"&offset={offset}&limit=1000",
                headers={**self.headers, "Prefer": ""}, timeout=15)
            batch = resp.json()
            if not batch: break
            all_br_mfn.extend(batch)
            offset += len(batch)
            if len(batch) < 1000: break

        mfn_map = {m["commoditycode"]: m for m in all_br_mfn}

        comm, mfn, vat = [], [], []
        for br in all_br:
            code = br["commoditycode"]
            comm.append({
                "commoditycode": code, "countrycode": "AR",
                "subheadingcode": br["subheadingcode"],
                "hsversion": "HS 2022",
                "nationaldescription": br["nationaldescription"],
                "codelength": "8-digit", "isactive": True,
            })
            br_mfn = mfn_map.get(code, {})
            rate = br_mfn.get("appliedmfnrate", 0)
            mfn.append({
                "commoditycode": code, "countrycode": "AR",
                "ratecategory": "APPLIED", "dutybasistype": "AD_VALOREM",
                "appliedmfnrate": rate,
                "dutyexpression": br_mfn.get("dutyexpression", f"{rate}%"),
                "valuationbasis": "CIF",
                "effectivefrom": self.today, "effectiveto": None,
            })
            # Argentina IVA = 21%
            vat.append({
                "commoditycode": code, "countrycode": "AR",
                "taxtype": "VAT", "taxcategory": "STANDARD",
                "rate": 21.0,
                "vatbasis": "CUSTOMS_VALUE_PLUS_DUTY",
                "postponedaccounting": False, "reliefavailable": False,
                "effectivefrom": self.today, "effectiveto": None,
                "notes": None,
            })

        c = self._upsert("commodity_code", comm)
        m = self._upsert("mfn_rate", mfn)
        v = self._upsert("vat_rate", vat)
        self._update_freshness("AR", c, "MERCOSUR TEC (copied from BR) + IVA 21%")
        logger.info("AR: %d codes, %d MFN rates, %d VAT rates", c, m, v)
        return c


def main():
    parser = argparse.ArgumentParser(description="Bulk country tariff loader")
    parser.add_argument("--country", help="Country code: UY, CL, AE, AR, or ALL")
    parser.add_argument("--all", action="store_true", help="Load all available countries")
    args = parser.parse_args()

    supabase_url = os.environ.get("SUPABASE_URL")
    supabase_key = os.environ.get("SUPABASE_SERVICE_KEY")
    if not supabase_url or not supabase_key:
        logger.error("Missing SUPABASE_URL or SUPABASE_SERVICE_KEY")
        sys.exit(1)

    loader = BulkCountryLoader(supabase_url, supabase_key)

    country = (args.country or "").upper()

    if country == "UY" or args.all:
        loader.load_uruguay()
    if country == "CL" or args.all:
        loader.load_chile()
    if country in ("AE", "SA", "OM", "GCC") or args.all:
        loader.load_gcc()
    if country == "AR" or args.all:
        loader.load_argentina_from_brazil()

    if not country and not args.all:
        logger.error("Specify --country UY|CL|AE|AR or --all")


if __name__ == "__main__":
    main()
